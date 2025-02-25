import pandas as pd
import talib as ta
import numpy as np
import pandas_ta as pta  # For SuperTrend calculation
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Configuration parameters
INITIAL_BALANCE = 10000.00
RISK_PER_TRADE = 0.01  # 1% risk per trade
SUPER_TREND_PERIOD = 10
SUPER_TREND_MULTIPLIER = 3
SMA_200_PERIOD = 200
ATR_PERIOD = 14
STOP_LOSS_CANDLES = 3  # Number of candles to look back for stop loss

# Load data
data_1h = pd.read_csv('BTCUSDT_1h.csv', parse_dates=['timestamp'])
data_5m = pd.read_csv('BTCUSDT_5m.csv', parse_dates=['timestamp'])

# Calculate indicators on 1h timeframe
# SuperTrend
supertrend_1h = pta.supertrend(data_1h['high'], data_1h['low'], data_1h['close'], length=SUPER_TREND_PERIOD, multiplier=SUPER_TREND_MULTIPLIER)
supertrend_col = [col for col in supertrend_1h.columns if col.startswith('SUPERT_') and str(SUPER_TREND_PERIOD) in col and str(SUPER_TREND_MULTIPLIER) in col][0]
supertrend_direction_col = [col for col in supertrend_1h.columns if col.startswith('SUPERTd_') and str(SUPER_TREND_PERIOD) in col and str(SUPER_TREND_MULTIPLIER) in col][0]
data_1h['SuperTrend'] = supertrend_1h[supertrend_col]  # SuperTrend line value
data_1h['SuperTrend_Direction'] = supertrend_1h[supertrend_direction_col]  # SuperTrend direction

# SMA
data_1h['SMA_200'] = ta.SMA(data_1h['close'], timeperiod=SMA_200_PERIOD)

# Determine trend on 1h timeframe (using SuperTrend and SMA_200 only)
data_1h['Trend'] = np.where(
    (data_1h['SuperTrend_Direction'] == 1) & 
    (data_1h['close'] > data_1h['SMA_200']), 'Up',
    np.where(
        (data_1h['SuperTrend_Direction'] == -1) & 
        (data_1h['close'] < data_1h['SMA_200']), 'Down', 'None'))

# Calculate indicators on 5m timeframe
data_5m['EMA5'] = ta.EMA(data_5m['close'], timeperiod=5)
data_5m['EMA8'] = ta.EMA(data_5m['close'], timeperiod=8)
data_5m['EMA13'] = ta.EMA(data_5m['close'], timeperiod=13)
data_5m['ATR'] = ta.ATR(data_5m['high'], data_5m['low'], data_5m['close'], timeperiod=ATR_PERIOD)

# Align 1h trend with 5m data
data_5m = pd.merge_asof(data_5m, data_1h[['timestamp', 'Trend']], on='timestamp', direction='backward')

# Initialize trading variables
balance = INITIAL_BALANCE
trades = []
in_position = False
position_type = None
entry_price = None
stop_loss = None
take_profit = None  # Simplified to a single take-profit level
position_size = None
entry_time = None
risk_amount = None

# Trading strategy simulation on 5m timeframe
for i in range(STOP_LOSS_CANDLES, len(data_5m)):
    if not in_position:
        current_close = data_5m['close'][i] if not pd.isna(data_5m['close'][i]) else 0
        ema5 = data_5m['EMA5'][i] if not pd.isna(data_5m['EMA5'][i]) else 0
        ema8 = data_5m['EMA8'][i] if not pd.isna(data_5m['EMA8'][i]) else 0
        ema13 = data_5m['EMA13'][i] if not pd.isna(data_5m['EMA13'][i]) else 0

        # Long entry conditions: All EMAs crossed up (price > 5 > 8 > 13) and close above EMAs, in an uptrend
        if (data_5m['Trend'][i] == 'Up' and 
            current_close > ema5 and ema5 > ema8 and ema8 > ema13 and 
            current_close > ema13):  # Ensure close is above the slowest EMA
            entry_price = data_5m['close'][i]
            # Set stop loss below the low of the last 3 candles, adjusted by 2x ATR
            last_3_lows = min(data_5m['low'][i - STOP_LOSS_CANDLES:i]) if not pd.isna(data_5m['low'][i - STOP_LOSS_CANDLES:i]).any() else data_5m['low'][i]
            atr_at_entry = data_5m['ATR'][i] if not pd.isna(data_5m['ATR'][i]) else 0
            stop_loss = last_3_lows - (2 * atr_at_entry)
            stop_distance = entry_price - stop_loss
            if stop_distance > 0:
                risk_amount = balance * RISK_PER_TRADE
                position_size = risk_amount / stop_distance
                # Take-profit at 1:3 risk-reward ratio
                take_profit = entry_price + (3 * stop_distance)
                in_position = True
                position_type = 'Long'
                entry_time = data_5m['timestamp'][i]

        # Short entry conditions: All EMAs crossed down (13 > 8 > 5 > price) and close below EMAs, in a downtrend
        elif (data_5m['Trend'][i] == 'Down' and 
              ema13 > ema8 and ema8 > ema5 and ema5 > current_close and 
              current_close < ema5):  # Ensure close is below the fastest EMA
            entry_price = data_5m['close'][i]
            # Set stop loss above the high of the last 3 candles, adjusted by 2x ATR
            last_3_highs = max(data_5m['high'][i - STOP_LOSS_CANDLES:i]) if not pd.isna(data_5m['high'][i - STOP_LOSS_CANDLES:i]).any() else data_5m['high'][i]
            atr_at_entry = data_5m['ATR'][i] if not pd.isna(data_5m['ATR'][i]) else 0
            stop_loss = last_3_highs + (2 * atr_at_entry)
            stop_distance = stop_loss - entry_price
            if stop_distance > 0:
                risk_amount = balance * RISK_PER_TRADE
                position_size = risk_amount / stop_distance
                # Take-profit at 1:3 risk-reward ratio
                take_profit = entry_price - (3 * stop_distance)
                in_position = True
                position_type = 'Short'
                entry_time = data_5m['timestamp'][i]

    elif in_position:
        exit_price = None
        exit_reason = None

        if position_type == 'Long':
            # Check stop loss
            if data_5m['low'][i] <= stop_loss:
                exit_price = stop_loss
                exit_reason = 'SL Triggered'
            # Check take-profit (1:3 risk-reward)
            elif data_5m['high'][i] >= take_profit:
                exit_price = take_profit
                exit_reason = 'TP Triggered'

        elif position_type == 'Short':
            # Check stop loss
            if data_5m['high'][i] >= stop_loss:
                exit_price = stop_loss
                exit_reason = 'SL Triggered'
            # Check take-profit (1:3 risk-reward)
            elif data_5m['low'][i] <= take_profit:
                exit_price = take_profit
                exit_reason = 'TP Triggered'

        # Process trade exit if triggered
        if exit_price is not None:
            # Calculate PnL
            if position_type == 'Long':
                pnl = (exit_price - entry_price) * position_size
            else:
                pnl = (entry_price - exit_price) * position_size
            balance += pnl
            
            # Record trade with additional entry indicators
            trades.append({
                'Position': position_type,
                'Entry Time': entry_time,
                'Exit Time': data_5m['timestamp'][i],
                'Entry Price': entry_price,
                'Stop Loss': stop_loss,
                'Take Profit': take_profit,  # Record single take-profit
                'Exit Price': exit_price,
                'Position Size': position_size,
                'Risk Amount': risk_amount,
                'PnL': pnl,
                'Balance After Trade': balance,
                'Exit Reason': exit_reason,
                'RSI at Entry': 0,  # Removed RSI, so set to 0 (or remove if not needed)
                'ATR at Entry': data_5m['ATR'][i] if not pd.isna(data_5m['ATR'][i]) else 0
            })
            in_position = False
            position_type = None
            entry_price = None
            stop_loss = None
            take_profit = None
            position_size = None
            entry_time = None
            risk_amount = None

# Create trades DataFrame
trades_df = pd.DataFrame(trades)

# Calculate summary statistics with win rate as decimal (for internal calculation)
# Ensure PnL is numeric and handle any non-numeric values
trades_df['PnL'] = pd.to_numeric(trades_df['PnL'], errors='coerce')
winning_trades = len(trades_df[trades_df['PnL'] > 0])
total_trades = len(trades_df)
win_rate = winning_trades / total_trades if total_trades > 0 else 0  # Decimal, not percentage

summary = {
    'Initial Balance': INITIAL_BALANCE,
    'Final Balance': balance,
    'Total Return': balance - INITIAL_BALANCE,
    'Total Trades': total_trades,
    'Win Rate': win_rate  # Decimal value (e.g., 0.2584)
}
summary_df = pd.DataFrame([summary])

# Write to Excel with formatting
with pd.ExcelWriter('backtest_results.xlsx', engine='openpyxl') as writer:
    trades_df.to_excel(writer, sheet_name='Trades', index=False)
    summary_df.to_excel(writer, sheet_name='Summary', index=False)

    # Get workbook and worksheets
    wb = writer.book
    ws_trades = wb['Trades']
    ws_summary = wb['Summary']

    # Adjust column widths
    for worksheet in [ws_trades, ws_summary]:
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

    # Format Trades sheet
    currency_format = '$#,##0.00'
    btc_format = '0.00000000'
    datetime_format = 'yyyy-mm-dd hh:mm:ss'
    number_format = '0.00'  # For RSI and ATR values

    for row in ws_trades.iter_rows(min_row=2, max_row=ws_trades.max_row):
        for cell in row:
            col_idx = cell.column
            # Entry Time and Exit Time
            if col_idx in [2, 3]:
                cell.number_format = datetime_format
            # Entry Price, Stop Loss, Take Profit, Exit Price, Risk Amount, PnL, Balance
            elif col_idx in [4, 5, 6, 7, 9, 10, 11]:
                cell.number_format = currency_format
            # Position Size
            elif col_idx == 8:
                cell.number_format = btc_format
            # Exit Reason, RSI at Entry, ATR at Entry
            elif col_idx in [12, 13, 14]:
                cell.number_format = number_format if col_idx in [13, 14] else '@'  # Text format for Exit Reason

    # Format Summary sheet
    percentage_format = '0.00%'  # For win rate as percentage
    
    for row in ws_summary.iter_rows(min_row=2, max_row=2):
        for cell in row:
            col_idx = cell.column
            # Initial Balance, Final Balance, Total Return
            if col_idx in [1, 2, 3]:
                cell.number_format = currency_format
            # Win Rate
            elif col_idx == 5:
                cell.number_format = percentage_format  # Percentage format (e.g., 0.00%)